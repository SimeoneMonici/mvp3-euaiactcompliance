import streamlit as st
import pandas as pd
import json
from questions_config import questions, per_system_qs, extra_qs
from utils import render_question
from gap_calculator import calculate_results, check_condition
import openpyxl
from io import BytesIO

st.title("AI Compliance Analyzer - Web App v6.70")

print("Avvio app...")

# Inizializzazione stato
if 'answers' not in st.session_state:
    st.session_state.answers = {q["id"]: "" for q in questions}
    print("Inizializzato st.session_state.answers")
if 'system_answers' not in st.session_state:
    st.session_state.system_answers = [{} for _ in range(1)]
    print("Inizializzato st.session_state.system_answers")
if 'step' not in st.session_state:
    st.session_state.step = 0
    print("Inizializzato st.session_state.step")
if 'num_systems' not in st.session_state:
    st.session_state.num_systems = 1
    print("Inizializzato st.session_state.num_systems")
if 'current_system' not in st.session_state:
    st.session_state.current_system = 0
    print("Inizializzato st.session_state.current_system")

print(f"Stato iniziale: step={st.session_state.step}, num_systems={st.session_state.num_systems}, current_system={st.session_state.current_system}")

# Sidebar
with st.sidebar:
    print("Inizio rendering sidebar")
    st.header("Navigazione")
    if st.button("Reinizializza"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.session_state.answers = {q["id"]: "" for q in questions}
        st.session_state.system_answers = [{} for _ in range(1)]
        st.session_state.step = 0
        st.session_state.num_systems = 1
        st.session_state.current_system = 0
        print("Reinizializzazione completata")
        st.rerun()
    if st.session_state.get('step', 0) > 0:
        if st.button("Indietro"):
            st.session_state.step -= 1
            st.session_state.current_system = 0
            print(f"Torna a step {st.session_state.step}")
            st.rerun()
    st.write(f"Step: {st.session_state.get('step', 0)} / 3")
    st.subheader("Risposte Chiave")
    if 'answers' in st.session_state:
        st.write(f"Settore: {st.session_state.answers.get('q1_1', 'Non specificato')}")
        st.write(f"Dimensione: {st.session_state.answers.get('q1_2', 'Non specificato')}")
    if 'system_answers' in st.session_state:
        for i, sys in enumerate(st.session_state.system_answers):
            ruolo = sys.get('q2_4', 'Non specificato')
            st.write(f"Sistema {i+1}: {sys.get('q2_1', 'Non specificato')} (Ruolo: {ruolo}, Caso: {sys.get('q2_5', 'Non specificato')}, Rischio: {sys.get('q2_8', 'Non specificato')})")

# Step 0: Profilo Aziendale
if st.session_state.step == 0:
    print("Esecuzione Step 0: Profilo Aziendale")
    st.header("Profilo Aziendale")
    placeholders = {q["id"]: st.empty() for q in questions}
    for q in questions:
        if not q.get('condition') or all(st.session_state.answers.get(cond_id, "") == cond_val for cond_id, cond_val in q.get('condition', {}).items()):
            with placeholders[q["id"]]:
                st.markdown(f"{q['question']}")
                current_answer = render_question(q, q["id"], st.session_state.answers.get(q["id"], ""))
                st.session_state.answers[q["id"]] = current_answer
                print(f"Q {q['id']} risposta: {current_answer}")
                if q["id"] == "q1_1" and current_answer in ["Difesa/Militare", "Sicurezza Nazionale"]:
                    st.warning("L'AI Act non si applica a questo settore per scopi militari o di sicurezza nazionale (Emendamento 808, Art. 2). Contatta un consulente per normative nazionali.")
    with st.sidebar:
        st.number_input("Numero sistemi IA", min_value=1, value=st.session_state.num_systems, key="num_systems_input", on_change=lambda: st.session_state.update({'num_systems': st.session_state.num_systems_input}) or st.rerun())
        st.session_state.num_systems = st.session_state.num_systems_input
        st.session_state.system_answers = st.session_state.system_answers[:st.session_state.num_systems] + [{} for _ in range(max(0, st.session_state.num_systems - len(st.session_state.system_answers)))]
        if st.button("Avanti a Inventory Sistemi"):
            st.session_state.step = 1
            print("Avanza a Step 1")
            st.rerun()

# Step 1: Inventory Sistemi (base)
elif st.session_state.step == 1:
    print(f"Esecuzione Step 1: Inventory Sistemi ({st.session_state.current_system + 1}/{st.session_state.num_systems})")
    st.header(f"Inventory Sistemi IA ({st.session_state.current_system + 1}/{st.session_state.num_systems})")
    sys_ans = st.session_state.system_answers[st.session_state.current_system]
    with st.form(key=f"system_form_{st.session_state.current_system}"):
        placeholders = {q["id"]: st.empty() for q in per_system_qs}
        for q in per_system_qs:
            condition_met = check_condition(q.get('condition', {}), {**st.session_state.answers, **sys_ans})
            if condition_met:
                with placeholders[q["id"]]:
                    st.markdown(f"{q['question']}")
                    sys_ans[q['id']] = render_question(q, q["id"], sys_ans.get(q["id"], ""))
                    print(f"Q {q['id']} risposta: {sys_ans[q['id']]}")
        if st.form_submit_button("Salva e Avanza"):
            st.session_state.step = 2
            print("Avanza a Step 2")
            st.rerun()
    with st.sidebar:
        if st.button("Aggiungi Sistema"):
            st.session_state.num_systems += 1
            st.session_state.system_answers.append({})
            print(f"Aggiunto sistema, totale: {st.session_state.num_systems}")
            st.rerun()

# Step 2: Domande Extra
elif st.session_state.step == 2:
    print(f"Esecuzione Step 2: Domande Extra Parametrizzate (Sistema {st.session_state.current_system + 1})")
    st.header("Domande Extra Parametrizzate")
    sys_ans = st.session_state.system_answers[st.session_state.current_system]
    with st.form(key=f"extra_form_{st.session_state.current_system}"):
        st.write(f"Basato su tue risposte (Settore: {st.session_state.answers.get('q1_1', 'Non specificato')}, Ruolo: {sys_ans.get('q2_4', 'Non specificato')}, Rischio: {sys_ans.get('q2_8', 'Non specificato')}, Caso d'uso: {sys_ans.get('q2_5', 'Non specificato')}), vedrai domande raggruppate.")
        seen_questions = set()
        displayed_questions = {}
        all_extra_questions = []
        for cat_val in extra_qs.values():
            if isinstance(cat_val, dict):
                for sub_cat, sub_qs in cat_val.items():
                    all_extra_questions.extend(sub_qs if isinstance(sub_qs, list) else [sub_qs])
            else:  # Gestione di categorie come 'esclusioni' che sono liste
                all_extra_questions.extend(cat_val)
        # Raggruppamento per categoria con priorità: Settore > Ruolo > Rischio > Caso d'Uso > Esclusioni
        categories = ['settore', 'ruolo', 'rischio', 'caso_uso', 'esclusioni']
        for cat in categories:
            if cat in extra_qs:
                if isinstance(extra_qs[cat], dict):
                    for sub_cat, sub_qs in extra_qs[cat].items():
                        for q in sub_qs if isinstance(sub_qs, list) else [sub_qs]:
                            q_id = q['id']
                            cond = q.get('condition', {})
                            if check_condition(cond, {**st.session_state.answers, **sys_ans}) and q_id not in seen_questions:
                                displayed_questions.setdefault(cat, []).append(q)
                                seen_questions.add(q_id)
                else:  # Caso lista (es. esclusioni)
                    for q in extra_qs[cat]:
                        q_id = q['id']
                        cond = q.get('condition', {})
                        if check_condition(cond, {**st.session_state.answers, **sys_ans}) and q_id not in seen_questions:
                            displayed_questions.setdefault(cat, []).append(q)
                            seen_questions.add(q_id)
        # Renderizza con expander
        if 'settore' in displayed_questions and st.session_state.answers.get('q1_1'):
            with st.expander(f"Domande per Settore: {st.session_state.answers['q1_1']} ({', '.join([ref['source'] for q in displayed_questions['settore'] for ref in q.get('references', []) if ref] or ['N/A'])}) - Perché? Requisiti settoriali"):
                for q in displayed_questions.get('settore', []):
                    st.markdown(f"{q['question']}")
                    sys_ans[q['id']] = render_question(q, f"{q['id']}_settore_{q_id}", sys_ans.get(q['id'], ""))  # Chiave unica
                    print(f"Q {q['id']} risposta: {sys_ans[q['id']]}")
                    if q['id'] == 'q2_38_sanita' and q not in displayed_questions.get('settore', []):  # Forza q2_38_sanita
                        displayed_questions['settore'].append(q)
        if 'ruolo' in displayed_questions and sys_ans.get('q2_4'):
            with st.expander(f"Domande per Ruolo: {sys_ans['q2_4']} ({', '.join([ref['source'] for q in displayed_questions['ruolo'] for ref in q.get('references', []) if ref] or ['N/A'])}) - Perché? Obblighi specifici per il tuo ruolo"):
                for q in displayed_questions.get('ruolo', []):
                    st.markdown(f"{q['question']}")
                    sys_ans[q['id']] = render_question(q, f"{q['id']}_ruolo_{q_id}", sys_ans.get(q['id'], ""))  # Chiave unica
                    print(f"Q {q['id']} risposta: {sys_ans[q['id']]}")
        if 'rischio' in displayed_questions and sys_ans.get('q2_8'):
            with st.expander(f"Domande per Rischio: {sys_ans['q2_8']} ({', '.join([ref['source'] for q in displayed_questions['rischio'] for ref in q.get('references', []) if ref] or ['N/A'])}) - Perché? Impatti legati al livello di rischio"):
                for q in displayed_questions.get('rischio', []):
                    st.markdown(f"{q['question']}")
                    sys_ans[q['id']] = render_question(q, f"{q['id']}_rischio_{q_id}", sys_ans.get(q['id'], ""))  # Chiave unica
                    print(f"Q {q['id']} risposta: {sys_ans[q['id']]}")
        if 'caso_uso' in displayed_questions and sys_ans.get('q2_5'):
            with st.expander(f"Domande per Caso d'Uso: {sys_ans['q2_5']} ({', '.join([ref['source'] for q in displayed_questions['caso_uso'] for ref in q.get('references', []) if ref] or ['N/A'])}) - Perché? Specifiche del tuo utilizzo"):
                for q in displayed_questions.get('caso_uso', []):
                    st.markdown(f"{q['question']}")
                    sys_ans[q['id']] = render_question(q, f"{q['id']}_caso_uso_{q_id}", sys_ans.get(q['id'], ""))  # Chiave unica
                    print(f"Q {q['id']} risposta: {sys_ans[q['id']]}")
        if 'esclusioni' in displayed_questions:
            with st.expander(f"Domande per Esclusioni ({', '.join([ref['source'] for q in displayed_questions['esclusioni'] for ref in q.get('references', []) if ref] or ['N/A'])}) - Perché? Verifica di eventuali esenzioni"):
                for q in displayed_questions.get('esclusioni', []):
                    st.markdown(f"{q['question']}")
                    sys_ans[q['id']] = render_question(q, f"{q['id']}_esclusioni_{q_id}", sys_ans.get(q['id'], ""))  # Chiave unica
                    print(f"Q {q['id']} risposta: {sys_ans[q['id']]}")
        # Debug in fondo
        with st.expander("Log Debug"):
            for q in all_extra_questions:
                q_id = q['id']
                cond = q.get('condition', {})
                if check_condition(cond, {**st.session_state.answers, **sys_ans}):
                    if q_id in seen_questions:
                        if isinstance(cond, dict):
                            st.write(f"Debug: Mostrando {q_id} perché condition {cond} ok con responses { {k: v for k, v in {**st.session_state.answers, **sys_ans}.items() if k in cond} }")
                        else:
                            st.write(f"Debug: Mostrando {q_id} perché condition {cond} (semplice booleano) ok")
                    else:
                        st.write(f"Debug: Escluso {q_id} perché condition {cond} non soddisfatta o già visto")
        if st.form_submit_button("Salva e Avanza"):
            if st.session_state.current_system < st.session_state.num_systems - 1:
                st.session_state.current_system += 1
            else:
                st.session_state.step = 3
            print(f"Avanza a step {st.session_state.step}, sistema {st.session_state.current_system}")
            st.rerun()
    with st.sidebar:
        if st.button("Avanti Sistema"):
            st.session_state.current_system += 1
            if st.session_state.current_system >= st.session_state.num_systems:
                st.session_state.step = 3
            print(f"Avanti sistema a {st.session_state.current_system}")
            st.rerun()
        if st.button("Aggiungi Sistema"):
            st.session_state.num_systems += 1
            st.session_state.system_answers.append({})
            print(f"Aggiunto sistema, totale: {st.session_state.num_systems}")
            st.rerun()

# Step 3: Risultati
elif st.session_state.step == 3:
    print("Esecuzione Step 3: Risultati")
    st.header("Risultati")
    result = calculate_results(st.session_state)
    general_pct, general_gaps, system_scores, system_gaps_list = result
    print(f"Risultati: general_pct={general_pct}, system_scores={system_scores}")
    sector = st.session_state.answers.get('q1_1', '')
    disclaimer_shown = False
    for sys_ans in st.session_state.system_answers:
        q2_43_response = sys_ans.get('q2_43', '')
        if sector in ["Difesa/Militare", "Sicurezza Nazionale"] and q2_43_response != 'Sì':
            st.warning("Nota: L'AI Act non si applica a questo settore per scopi militari o di sicurezza nazionale (Art. 2) a meno che non sia escluso esplicitamente (q2_43).")
            disclaimer_shown = True
    if disclaimer_shown:
        st.info("Analisi terminata per esclusione settoriale.")
    else:
        st.write(f"Totale Conformità Generale: {general_pct:.2f}%")
        for i, sys_pct in enumerate(system_scores):
            st.write(f"Sistema {i+1}: {sys_pct:.2f}%")
        avg_sys = sum(system_scores) / len(system_scores) if system_scores else 0
        st.write(f"Media Sistemi: {avg_sys:.2f}%")
        st.header("Censimento Sistemi IA")
        censimento_data = []
        for i, sys_ans in enumerate(st.session_state.system_answers):
            censimento_data.append({
                "Sistema": i+1,
                "Nome": sys_ans.get("q2_1", "Non specificato"),
                "Descrizione": sys_ans.get("q2_2", "Non specificato"),
                "Funzione": sys_ans.get("q2_3", "Non specificato"),
                "Ruolo": sys_ans.get("q2_4", "Non specificato"),
                "Caso d'uso": sys_ans.get("q2_5", "Non specificato"),
                "Rischio": sys_ans.get("q2_8", "N/A"),
                "Punteggio Conformità": f"{system_scores[i]:.2f}%" if i < len(system_scores) else "N/A"
            })
        df_censimento = pd.DataFrame(censimento_data)
        st.table(df_censimento)
        st.header("Roadmap di Adeguamento")
        all_gaps = general_gaps + [gap for sys_gaps in system_gaps_list for gap in sys_gaps]
        if all_gaps:
            df = pd.DataFrame(all_gaps)
            st.table(df[["Gap", "Risposta", "Azione", "Priorità", "Tempistica", "Risk"]])
            csv = df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8')
            st.download_button(label="Scarica Roadmap CSV", data=csv, file_name="roadmap_gaps.csv", mime="text/csv")
        else:
            st.warning("Nessun gap rilevato.")
        # Riepilogo Risposte
        st.header("Riepilogo Risposte")
        riepilogo_data = []
        for q_id, ans in st.session_state.answers.items():
            matching_question = next((q for q in questions if q['id'] == q_id), None)
            riepilogo_data.append({"Domanda": matching_question['question'] if matching_question else q_id, "Risposta": str(ans)})
        for sys_idx, sys_ans in enumerate(st.session_state.system_answers):
            for q_id, ans in sys_ans.items():
                all_questions = per_system_qs + [q for cat in extra_qs.values() for sub in (cat.values() if isinstance(cat, dict) else [cat]) for q in (sub if isinstance(sub, list) else [sub])]
                matching_question = next((q for q in all_questions if q['id'] == q_id), None)
                riepilogo_data.append({"Domanda": matching_question['question'] if matching_question else q_id, "Risposta": str(ans), "Sistema": sys_idx + 1})
        df_riepilogo = pd.DataFrame(riepilogo_data)
        st.table(df_riepilogo)
        # Debug in fondo
        with st.expander("Log Debug"):
            st.write("Nessun log di debug attivo in questa sezione.")
        # Esportazione Excel
        wb = openpyxl.Workbook()
        ws_censimento = wb.active
        ws_censimento.title = "Censimento"
        ws_censimento.append(["Sistema", "Nome", "Descrizione", "Funzione", "Ruolo", "Caso d'uso", "Rischio", "Punteggio Conformità"])
        for row in censimento_data:
            ws_censimento.append([row["Sistema"], row["Nome"], row["Descrizione"], row["Funzione"], row["Ruolo"], row["Caso d'uso"], row["Rischio"], row["Punteggio Conformità"]])
        ws_roadmap = wb.create_sheet("Roadmap")
        ws_roadmap.append(["Gap", "Risposta", "Azione", "Priorità", "Tempistica", "Risk"])
        for row in all_gaps:
            ws_roadmap.append([row["Gap"], row["Risposta"], row["Azione"], row["Priorità"], row["Tempistica"], row["Risk"]])
        ws_riepilogo = wb.create_sheet("Riepilogo")
        ws_riepilogo.append(["Metrica", "Valore"])
        ws_riepilogo.append(["Totale Conformità Generale", f"{general_pct:.2f}%"])
        ws_riepilogo.append(["Media Sistemi", f"{avg_sys:.2f}%"])
        for i, score in enumerate(system_scores):
            ws_riepilogo.append([f"Sistema {i+1} Conformità", f"{score:.2f}%"])
        output = BytesIO()
        wb.save(output)
        st.download_button(label="Scarica Report Excel", data=output.getvalue(), file_name="ai_compliance_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")