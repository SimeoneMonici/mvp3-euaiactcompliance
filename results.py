import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import json  # <--- AGGIUNTO
from questions_config import questions, per_system_qs, extra_qs
from gap_calculator import calculate_results

def load_recommendations(role):
    role_map = {
        "Sviluppatore": "recommendations_sviluppatore.json",
        "Utilizzatore": "recommendations_user.json",
        "Importatore": "recommendations_importer.json",
        "Distributore": "recommendations_distributore.json"
    }
    path = role_map.get(role, "recommendations_general.json")
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return {r["question_id"]: r for r in data["recommendations"]}
    except FileNotFoundError:
        st.error(f"File recommendations mancante: {path}")
        return {}

def generate_excel():
    general_pct, general_gaps, system_scores, system_gaps = calculate_results(st.session_state)
    wb = openpyxl.Workbook()
    
    # FOGLIO DOMANDE
    ws_dom = wb.active
    ws_dom.title = "Domande"
    headers = ["Sezione", "Domanda", "Tipo", "Riferimento Normativo", "Note", "Score Weight", "Risk Flag"]
    for col, header in enumerate(headers, 1):
        ws_dom.cell(1, col, header).font = Font(bold=True)
    row = 2
    for q in questions:
        ws_dom.cell(row, 1, q.get("expander", q.get("section", "N/A")))
        ws_dom.cell(row, 2, q["question"])
        ws_dom.cell(row, 3, q.get("type"))
        ws_dom.cell(row, 4, q["ref"])
        ws_dom.cell(row, 5, q["notes"])
        ws_dom.cell(row, 6, q.get("score_weight", 1))
        ws_dom.cell(row, 7, q.get("risk_flag", "N/A"))
        row += 1
    for cat, sub_qs in extra_qs.items():
        if isinstance(sub_qs, dict):
            for subcat, qs in sub_qs.items():
                ws_dom.cell(row, 1, f"Extra {cat} - {subcat}")
                row += 1
                for q in qs:
                    ws_dom.cell(row, 2, q["question"])
                    ws_dom.cell(row, 3, q.get("type"))
                    ws_dom.cell(row, 4, q.get("ref", ""))
                    ws_dom.cell(row, 5, q.get("notes", ""))
                    ws_dom.cell(row, 6, q.get("score_weight", 1))
                    ws_dom.cell(row, 7, q.get("risk_flag", "N/A"))
                    row += 1

    # FOGLIO RISPOSTE
    ws_resp = wb.create_sheet("Risposte")
    ws_resp.cell(1, 1, "Domanda")
    ws_resp.cell(1, 2, "Risposta")
    for i, q in enumerate(questions, 2):
        ws_resp.cell(i, 1, q["question"])
        ws_resp.cell(i, 2, str(st.session_state.answers.get(q["id"], "")))

    # FOGLIO INVENTORY
    ws_inv = wb.create_sheet("Inventory_Sistemi_IA")
    inv_headers = ["ID", "Nome", "Descrizione", "Funzione"] + [q["question"] for q in per_system_qs] + ["Punteggio (%)"]
    for col, header in enumerate(inv_headers, 1):
        ws_inv.cell(1, col, header).font = Font(bold=True)
    for i, sys_ans in enumerate(st.session_state.system_answers, 2):
        ws_inv.cell(i, 1, i-1)
        ws_inv.cell(i, 2, sys_ans.get("q2_1", ""))
        ws_inv.cell(i, 3, sys_ans.get("q2_2", ""))
        ws_inv.cell(i, 4, sys_ans.get("q2_3", ""))
        col_idx = 5
        for q in per_system_qs:
            ws_inv.cell(i, col_idx, str(sys_ans.get(q["id"], "")))
            col_idx += 1
        ws_inv.cell(i, col_idx, f"{system_scores[i-2]:.1f}" if i-2 < len(system_scores) else "0.0")

    # FOGLIO ROADMAP
    ws_road = wb.create_sheet("Roadmap")
    road_headers = ["Sistema", "Gap", "Risposta", "Azione", "Priorità", "Tempistica", "Risk"]
    for col, header in enumerate(road_headers, 1):
        ws_road.cell(1, col, header).font = Font(bold=True)
    row = 2
    # GAP GENERALI
    for gid, g in general_gaps.items():
        if gid.startswith('q1_'):
            q_id = gid
            role = "Generale"
            recs = load_recommendations(role)
            r = recs.get(q_id, {})
            ws_road.cell(row, 1, "Generale")
            ws_road.cell(row, 2, f"{gid}: {g['question']}")
            ws_road.cell(row, 3, g["response"])
            ws_road.cell(row, 4, r.get("recommendation", "Aggiungi in recommendations_general.json"))
            ws_road.cell(row, 5, r.get("priority", "Alta"))
            ws_road.cell(row, 6, r.get("timeline", "3 mesi"))
            ws_road.cell(row, 7, g["risk_flag"])
            row += 1
    # GAP SISTEMI
    for sys_gaps in system_gaps:
        for gap in sys_gaps["Gaps"]:
            q_id = gap["Gap"].split(":")[0].strip()
            role = next((s["q2_4"] for s in st.session_state.system_answers if s.get("q2_1", "").lower() in sys_gaps["system_name"].lower()), "")
            recs = load_recommendations(role)
            r = recs.get(q_id, {})
            ws_road.cell(row, 1, sys_gaps["system_name"])
            ws_road.cell(row, 2, gap["Gap"])
            ws_road.cell(row, 3, gap["Risposta"])
            ws_road.cell(row, 4, r.get("recommendation", "Aggiungi in recommendations_{role}.json"))
            ws_road.cell(row, 5, r.get("priority", "Alta"))
            ws_road.cell(row, 6, r.get("timeline", "3 mesi"))
            ws_road.cell(row, 7, gap["Risk"])
            row += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_roadmap_csv():
    general_pct, general_gaps, system_scores, system_gaps = calculate_results(st.session_state)
    df_list = []
    # GAP GENERALI
    for gid, g in general_gaps.items():
        if gid.startswith('q1_'):
            q_id = gid
            role = "Generale"
            recs = load_recommendations(role)
            r = recs.get(q_id, {})
            df_list.append({
                "Sistema": "Generale",
                "Gap": f"{gid}: {g['question']}",
                "Risposta": g["response"],
                "Azione": r.get("recommendation", "Aggiungi in recommendations_general.json"),
                "Priorità": r.get("priority", "Alta"),
                "Tempistica": r.get("timeline", "3 mesi"),
                "Risk": g["risk_flag"]
            })
    # GAP SISTEMI
    for sys_gaps in system_gaps:
        for gap in sys_gaps["Gaps"]:
            q_id = gap["Gap"].split(":")[0].strip()
            role = next((s["q2_4"] for s in st.session_state.system_answers if s.get("q2_1", "").lower() in sys_gaps["system_name"].lower()), "")
            recs = load_recommendations(role)
            r = recs.get(q_id, {})
            df_list.append({
                "Sistema": sys_gaps["system_name"],
                "Gap": gap["Gap"],
                "Risposta": gap["Risposta"],
                "Azione": r.get("recommendation", "Aggiungi in recommendations_{role}.json"),
                "Priorità": r.get("priority", "Alta"),
                "Tempistica": r.get("timeline", "3 mesi"),
                "Risk": gap["Risk"]
            })
    df = pd.DataFrame(df_list)
    output = BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    return output.getvalue()

def display_results():
    st.write("**AI Compliance v15.9 - Risultati**")
    try:
        general_pct, general_gaps, system_scores, system_gaps = calculate_results(st.session_state)
        st.metric("Conformità Generale", f"{general_pct:.1f}%")
        for i, score in enumerate(system_scores):
            name = st.session_state.system_answers[i].get("q2_1", f"Sistema {i+1}")
            st.metric(f"{name}", f"{score:.1f}%")
        st.header("Roadmap di Adeguamento")
        df_list = []
        # GAP GENERALI
        for gid, g in general_gaps.items():
            if gid.startswith('q1_'):
                q_id = gid
                role = "Generale"
                recs = load_recommendations(role)
                r = recs.get(q_id, {})
                df_list.append({
                    "Sistema": "Generale",
                    "Gap": f"{gid}: {g['question']}",
                    "Risposta": g["response"],
                    "Azione": r.get("recommendation", "Aggiungi in recommendations_general.json"),
                    "Priorità": r.get("priority", "Alta"),
                    "Tempistica": r.get("timeline", "3 mesi"),
                    "Risk": g["risk_flag"]
                })
        # GAP SISTEMI
        for sys_gaps in system_gaps:
            for gap in sys_gaps["Gaps"]:
                q_id = gap["Gap"].split(":")[0].strip()
                role = next((s["q2_4"] for s in st.session_state.system_answers if s.get("q2_1", "").lower() in sys_gaps["system_name"].lower()), "")
                recs = load_recommendations(role)
                r = recs.get(q_id, {})
                df_list.append({
                    "Sistema": sys_gaps["system_name"],
                    "Gap": gap["Gap"],
                    "Risposta": gap["Risposta"],
                    "Azione": r.get("recommendation", "Aggiungi in recommendations_{role}.json"),
                    "Priorità": r.get("priority", "Alta"),
                    "Tempistica": r.get("timeline", "3 mesi"),
                    "Risk": gap["Risk"]
                })
        if df_list:
            df = pd.DataFrame(df_list)
            st.dataframe(df, use_container_width=True)
        else:
            st.success("Nessun gap critico!")
    except Exception as e:
        st.error(f"Errore: {e}")
